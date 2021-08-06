# -*- coding: UTF-8 -*-

import os
import sys
import xlrd
import shutil
import logging
import argparse

from openpyxl import load_workbook
from openpyxl import Workbook


LOG = logging.getLogger(__name__)

__version__ = "1.0.0"
__author__ = ("Xingguo Zhang",)
__email__ = "invicoun@foxmail.com"
__all__ = []


def mv_special_str(lists):

    r = []

    for i in lists:
        try:
            i = i.replace("。", "").replace("\n", "").replace("\t", "").replace('"', "").replace('丨', "")
            i = i.replace(':', "").replace(',', "").replace(' ', "").replace('‘', "")
        except Exception as e:
            i = str(i)
        r.append(i)

    return r


def judgment_list(string, areas):

    r = False

    for i in areas:
        if i not in string:
            continue
        r = True
        break
    return r


def judgment_area(string1, string2):

    mn_areas = ["美年", "美", "年", "健康"]
    qs_areas = ["青山", "厂前", "工人", "石化", "九医", "钢都", "红钢城", "普仁", "冶金", "船厂", "白玉山"]
    mn_areas1 = ["欢乐","东沙花园", "华电", "嘉园", "鹏程", "蓝晶", "绿洲", "正堂", "华腾"]
    area = "希望组"

    if judgment_list(string1, mn_areas)  or judgment_list(string2, mn_areas):
        area = "美年"
    elif judgment_list(string1, qs_areas) or judgment_list(string1, qs_areas):
        area = "青山"
    elif judgment_list(string1, mn_areas1)  or judgment_list(string2, mn_areas1):
        area = "美年"
    else:
        pass

    return area


def write_to_excel(file, data):

    fo = Workbook()
    sheet = fo.active
    #sheet = fo['Sheet']
    #sheet.title = "sheet1"
    #sheet = fo.create_sheet("sheet1", index=0)

    for row_index, row_item in enumerate(data):
        for col_index, col_item in enumerate(row_item):
            sheet.cell(row=row_index+1, column= col_index+1, value=col_item)
        sheet.append(row_item)

    fo.save(file)
    LOG.info("Write the result to file %r" % file)
    LOG.info("The operation is successful, thank you!")

    return 0


def classified_xls(file, ofile, site1=10, site2=12):

    LOG.info("reading message from %r" % file)
    fh = load_workbook(file, data_only=True)
    table = fh.worksheets[0]
    LOG.info("dim: {} rows, {} columns.".format(table.max_row, table.max_column))

    r = []
    n = 0
    for line in table.values:
        n += 1
        line = mv_special_str(line)
        #print(line)
        area = judgment_area(line[site1-1], line[site2-1])
        line += [area]
        r.append(line)

    write_to_excel(ofile, r)

    return 0


def add_hlep_args(parser):

    parser.add_argument('input', metavar='FILE', type=str,
        help='Input file')
    parser.add_argument('-s1', '--site1', metavar='INT', type=int, default=10,
        help='Keyword position 1,default=11.')
    parser.add_argument('-s2', '--site2', metavar='INT', type=int, default=12,
        help='Keyword position 2,default=14.')
    parser.add_argument('-o', '--output', metavar='FILE', type=str, default="output.xlsx",
        help='Output file name,default=output.xlsx.')


    return parser


def main():

    logging.basicConfig(
        stream=sys.stderr,
        level=logging.INFO,
        format="[%(levelname)s] %(message)s"
    )
    parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter,
    description='''
name:
    classified_xlsx: Categorize excel files according to keywords
attention:
    classified_xlsx.py txt.xlsx
    classified_xlsx.py txt.xlsx -s1 10 -s2 12 -o txt_new.xlsx
version: %s
contact:  %s <%s>\
        ''' % (__version__, ' '.join(__author__), __email__))

    args = add_hlep_args(parser).parse_args()

    classified_xls(args.input, args.output, args.site1, args.site1)


if __name__ == "__main__":

    main()
